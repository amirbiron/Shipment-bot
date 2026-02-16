"""
שירות ייצוא דוחות — Excel מעוצב (openpyxl)

מספק פונקציות ליצירת קבצי XLSX מעוצבים עם כותרות, סיכומים, עמודות מותאמות,
ותמיכה בעברית (RTL).
"""
import io
from datetime import datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter


# ==================== קבועי עיצוב ====================

# צבעים
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(name="Arial", bold=True, size=14)
_SUBTITLE_FONT = Font(name="Arial", bold=False, size=10, color="666666")
_TOTAL_FONT = Font(name="Arial", bold=True, size=11)
_TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_CURRENCY_FORMAT = '#,##0.00 ₪'
_PERCENT_FORMAT = '0.00%'

# גבולות
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# יישור — RTL לעברית
_RTL_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_NUMBER_ALIGN = Alignment(horizontal="left", vertical="center")


def _auto_fit_columns(ws: Any) -> None:
    """התאמת רוחב עמודות אוטומטית לפי תוכן"""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
        # רוחב מינימלי 10, מקסימלי 40
        adjusted_width = min(max(max_length + 4, 10), 40)
        ws.column_dimensions[col_letter].width = adjusted_width


def _apply_header_style(ws: Any, row: int, col_count: int) -> None:
    """החלת עיצוב כותרות על שורה"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _RTL_ALIGN
        cell.border = _THIN_BORDER


def _apply_data_style(ws: Any, row: int, col_count: int) -> None:
    """החלת עיצוב נתונים על שורה"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = _RTL_ALIGN
        cell.border = _THIN_BORDER


def _apply_total_style(ws: Any, row: int, col_count: int) -> None:
    """החלת עיצוב שורת סיכום"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL
        cell.alignment = _RTL_ALIGN
        cell.border = _THIN_BORDER


def _write_title(ws: Any, title: str, subtitle: str, start_row: int = 1) -> int:
    """כתיבת כותרת ותת-כותרת, מחזיר את השורה הבאה"""
    ws.cell(row=start_row, column=1, value=title).font = _TITLE_FONT
    ws.cell(row=start_row + 1, column=1, value=subtitle).font = _SUBTITLE_FONT
    return start_row + 3  # שורה ריקה אחרי הכותרת


def _format_currency_cell(cell: Any) -> None:
    """עיצוב תא כמטבע"""
    cell.number_format = _CURRENCY_FORMAT
    cell.alignment = _NUMBER_ALIGN


# תווים מסוכנים — Excel מפרש אותם כנוסחה (Formula Injection / CSV Injection)
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_text(value: str) -> str:
    """ניטרול ערכי טקסט שעלולים להתפרש כנוסחה ב-Excel (Formula Injection).

    מוסיף גרש בודד (') כ-prefix — Excel יציג את הטקסט כרגיל בלי לפרש כנוסחה.
    """
    if isinstance(value, str) and value and value[0] in _FORMULA_PREFIXES:
        return f"'{value}"
    return value


# ==================== דוח גבייה — Excel ====================


def generate_collection_report_excel(
    items: list[dict[str, Any]],
    total_debt: float,
    cycle_start: str,
    cycle_end: str,
    station_name: str = "",
) -> bytes:
    """
    יצירת קובץ Excel מעוצב לדוח גבייה.

    Args:
        items: רשימת נהגים (driver_name, total_debt, charge_count)
        total_debt: סה"כ חוב
        cycle_start: תחילת מחזור (YYYY-MM-DD)
        cycle_end: סוף מחזור (YYYY-MM-DD)
        station_name: שם התחנה (אופציונלי)

    Returns:
        bytes — תוכן קובץ XLSX
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "דוח גבייה"
    ws.sheet_view.rightToLeft = True

    # כותרת
    title = "דוח גבייה"
    if station_name:
        title += f" — {_sanitize_text(station_name)}"
    subtitle = f"מחזור: {cycle_start} עד {cycle_end} | הופק: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    data_row = _write_title(ws, title, subtitle)

    # כותרות עמודות
    headers = ["שם נהג", "סה\"כ חוב (₪)", "מספר חיובים"]
    col_count = len(headers)
    for col, header in enumerate(headers, 1):
        ws.cell(row=data_row, column=col, value=header)
    _apply_header_style(ws, data_row, col_count)

    # נתונים
    for i, item in enumerate(items):
        row = data_row + 1 + i
        ws.cell(row=row, column=1, value=_sanitize_text(item["driver_name"]))
        debt_cell = ws.cell(row=row, column=2, value=float(item["total_debt"]))
        _format_currency_cell(debt_cell)
        ws.cell(row=row, column=3, value=item["charge_count"])
        _apply_data_style(ws, row, col_count)

    # שורת סיכום
    total_row = data_row + 1 + len(items)
    ws.cell(row=total_row, column=1, value="סה\"כ")
    total_cell = ws.cell(row=total_row, column=2, value=float(total_debt))
    _format_currency_cell(total_cell)
    ws.cell(row=total_row, column=3, value=sum(i["charge_count"] for i in items))
    _apply_total_style(ws, total_row, col_count)

    _auto_fit_columns(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ==================== דוח הכנסות — Excel ====================


def generate_revenue_report_excel(
    total_commissions: float,
    total_manual_charges: float,
    total_withdrawals: float,
    net_total: float,
    date_from: str,
    date_to: str,
    station_name: str = "",
) -> bytes:
    """
    יצירת קובץ Excel מעוצב לדוח הכנסות.

    Args:
        total_commissions: סה"כ עמלות
        total_manual_charges: סה"כ חיובים ידניים
        total_withdrawals: סה"כ משיכות
        net_total: נטו
        date_from: מתאריך
        date_to: עד תאריך
        station_name: שם התחנה (אופציונלי)

    Returns:
        bytes — תוכן קובץ XLSX
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "דוח הכנסות"
    ws.sheet_view.rightToLeft = True

    title = "דוח הכנסות"
    if station_name:
        title += f" — {_sanitize_text(station_name)}"
    subtitle = f"תקופה: {date_from} עד {date_to} | הופק: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    data_row = _write_title(ws, title, subtitle)

    # כותרות
    headers = ["סוג תנועה", "סכום (₪)"]
    col_count = len(headers)
    for col, header in enumerate(headers, 1):
        ws.cell(row=data_row, column=col, value=header)
    _apply_header_style(ws, data_row, col_count)

    # נתונים
    rows_data = [
        ("עמלות ממשלוחים", total_commissions),
        ("חיובים ידניים", total_manual_charges),
        ("משיכות", total_withdrawals),
    ]
    for i, (label, amount) in enumerate(rows_data):
        row = data_row + 1 + i
        ws.cell(row=row, column=1, value=label)
        amount_cell = ws.cell(row=row, column=2, value=float(amount))
        _format_currency_cell(amount_cell)
        _apply_data_style(ws, row, col_count)

    # שורת סיכום — נטו
    total_row = data_row + 1 + len(rows_data)
    ws.cell(row=total_row, column=1, value="נטו (עמלות + חיובים - משיכות)")
    net_cell = ws.cell(row=total_row, column=2, value=float(net_total))
    _format_currency_cell(net_cell)
    _apply_total_style(ws, total_row, col_count)

    _auto_fit_columns(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ==================== דוח רווח/הפסד — Excel ====================


def generate_profit_loss_excel(
    revenue_by_month: list[dict[str, Any]],
    date_from: str,
    date_to: str,
    station_name: str = "",
) -> bytes:
    """
    יצירת קובץ Excel מעוצב לדוח רווח/הפסד.

    Args:
        revenue_by_month: נתוני הכנסות והוצאות לכל חודש:
            - month: str (YYYY-MM)
            - commissions: float
            - manual_charges: float
            - withdrawals: float
            - net: float
        date_from: מתאריך
        date_to: עד תאריך
        station_name: שם התחנה

    Returns:
        bytes — תוכן קובץ XLSX
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "רווח והפסד"
    ws.sheet_view.rightToLeft = True

    title = "דוח רווח והפסד"
    if station_name:
        title += f" — {_sanitize_text(station_name)}"
    subtitle = f"תקופה: {date_from} עד {date_to} | הופק: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    data_row = _write_title(ws, title, subtitle)

    # כותרות
    headers = ["חודש", "עמלות (₪)", "חיובים ידניים (₪)", "משיכות (₪)", "נטו (₪)"]
    col_count = len(headers)
    for col, header in enumerate(headers, 1):
        ws.cell(row=data_row, column=col, value=header)
    _apply_header_style(ws, data_row, col_count)

    # נתונים חודשיים
    totals = {"commissions": 0.0, "manual_charges": 0.0, "withdrawals": 0.0, "net": 0.0}
    for i, month_data in enumerate(revenue_by_month):
        row = data_row + 1 + i
        ws.cell(row=row, column=1, value=month_data["month"])

        for j, key in enumerate(["commissions", "manual_charges", "withdrawals", "net"], 2):
            val = float(month_data[key])
            cell = ws.cell(row=row, column=j, value=val)
            _format_currency_cell(cell)
            totals[key] += val

        _apply_data_style(ws, row, col_count)

    # שורת סיכום
    total_row = data_row + 1 + len(revenue_by_month)
    ws.cell(row=total_row, column=1, value="סה\"כ")
    for j, key in enumerate(["commissions", "manual_charges", "withdrawals", "net"], 2):
        cell = ws.cell(row=total_row, column=j, value=totals[key])
        _format_currency_cell(cell)
    _apply_total_style(ws, total_row, col_count)

    _auto_fit_columns(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ==================== דוח חודשי מלא — Excel (מרובה גליונות) ====================


def generate_monthly_summary_excel(
    month: str,
    station_name: str,
    collection_items: list[dict[str, Any]],
    total_debt: float,
    revenue_data: dict[str, float],
    delivery_stats: dict[str, int],
) -> bytes:
    """
    יצירת דוח חודשי מלא עם מספר גליונות (sheets).

    Args:
        month: חודש הדוח (YYYY-MM)
        station_name: שם התחנה
        collection_items: נתוני גבייה (driver_name, total_debt, charge_count)
        total_debt: סה"כ חוב
        revenue_data: dict עם commissions, manual_charges, withdrawals, net
        delivery_stats: dict עם total, delivered, cancelled, open

    Returns:
        bytes — תוכן קובץ XLSX
    """
    wb = Workbook()

    # ==================== גליון 1: סיכום ====================
    ws_summary = wb.active
    ws_summary.title = "סיכום"
    ws_summary.sheet_view.rightToLeft = True

    subtitle = f"חודש: {month} | הופק: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    data_row = _write_title(ws_summary, f"דוח חודשי — {_sanitize_text(station_name)}", subtitle)

    # סטטיסטיקות משלוחים
    ws_summary.cell(row=data_row, column=1, value="סטטיסטיקות משלוחים").font = _TOTAL_FONT
    data_row += 1

    stats_data = [
        ("סה\"כ משלוחים", delivery_stats.get("total", 0)),
        ("נמסרו", delivery_stats.get("delivered", 0)),
        ("בוטלו", delivery_stats.get("cancelled", 0)),
        ("פתוחים", delivery_stats.get("open", 0)),
    ]
    headers = ["מדד", "ערך"]
    for col, h in enumerate(headers, 1):
        ws_summary.cell(row=data_row, column=col, value=h)
    _apply_header_style(ws_summary, data_row, 2)

    for i, (label, val) in enumerate(stats_data):
        row = data_row + 1 + i
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=val)
        _apply_data_style(ws_summary, row, 2)

    # סיכום פיננסי
    fin_row = data_row + len(stats_data) + 3
    ws_summary.cell(row=fin_row, column=1, value="סיכום פיננסי").font = _TOTAL_FONT
    fin_row += 1

    fin_headers = ["סוג", "סכום (₪)"]
    for col, h in enumerate(fin_headers, 1):
        ws_summary.cell(row=fin_row, column=col, value=h)
    _apply_header_style(ws_summary, fin_row, 2)

    fin_data = [
        ("עמלות", revenue_data.get("commissions", 0)),
        ("חיובים ידניים", revenue_data.get("manual_charges", 0)),
        ("משיכות", revenue_data.get("withdrawals", 0)),
        ("נטו", revenue_data.get("net", 0)),
    ]
    for i, (label, amount) in enumerate(fin_data):
        row = fin_row + 1 + i
        ws_summary.cell(row=row, column=1, value=label)
        cell = ws_summary.cell(row=row, column=2, value=float(amount))
        _format_currency_cell(cell)
        if label == "נטו":
            _apply_total_style(ws_summary, row, 2)
        else:
            _apply_data_style(ws_summary, row, 2)

    _auto_fit_columns(ws_summary)

    # ==================== גליון 2: גבייה ====================
    ws_collection = wb.create_sheet("גבייה")
    ws_collection.sheet_view.rightToLeft = True

    coll_row = _write_title(ws_collection, "דוח גבייה", f"חודש: {month}")

    coll_headers = ["שם נהג", "סה\"כ חוב (₪)", "מספר חיובים"]
    for col, h in enumerate(coll_headers, 1):
        ws_collection.cell(row=coll_row, column=col, value=h)
    _apply_header_style(ws_collection, coll_row, len(coll_headers))

    for i, item in enumerate(collection_items):
        row = coll_row + 1 + i
        ws_collection.cell(row=row, column=1, value=item["driver_name"])
        debt_cell = ws_collection.cell(row=row, column=2, value=float(item["total_debt"]))
        _format_currency_cell(debt_cell)
        ws_collection.cell(row=row, column=3, value=item["charge_count"])
        _apply_data_style(ws_collection, row, len(coll_headers))

    # שורת סיכום
    total_row = coll_row + 1 + len(collection_items)
    ws_collection.cell(row=total_row, column=1, value="סה\"כ")
    total_cell = ws_collection.cell(row=total_row, column=2, value=float(total_debt))
    _format_currency_cell(total_cell)
    ws_collection.cell(row=total_row, column=3, value=sum(i["charge_count"] for i in collection_items))
    _apply_total_style(ws_collection, total_row, len(coll_headers))

    _auto_fit_columns(ws_collection)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
