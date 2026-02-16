"""
בדיקות שירות ייצוא Excel
"""
import pytest
from openpyxl import load_workbook
import io

from app.domain.services.export_service import (
    _sanitize_text,
    generate_collection_report_excel,
    generate_revenue_report_excel,
    generate_profit_loss_excel,
    generate_monthly_summary_excel,
)


class TestSanitizeText:
    """בדיקות הגנה מ-Formula Injection"""

    @pytest.mark.unit
    def test_sanitize_formula_equals(self):
        """ערך שמתחיל ב-= מקבל גרש כ-prefix"""
        assert _sanitize_text("=SUM(A1:A10)") == "'=SUM(A1:A10)"

    @pytest.mark.unit
    def test_sanitize_formula_plus(self):
        """ערך שמתחיל ב-+ מקבל גרש"""
        assert _sanitize_text("+cmd|'/C calc'!A0") == "'+cmd|'/C calc'!A0"

    @pytest.mark.unit
    def test_sanitize_formula_minus(self):
        """ערך שמתחיל ב-- מקבל גרש"""
        assert _sanitize_text("-1+1") == "'-1+1"

    @pytest.mark.unit
    def test_sanitize_formula_at(self):
        """ערך שמתחיל ב-@ מקבל גרש"""
        assert _sanitize_text("@SUM(A1)") == "'@SUM(A1)"

    @pytest.mark.unit
    def test_sanitize_formula_tab(self):
        """ערך שמתחיל ב-tab מקבל גרש"""
        assert _sanitize_text("\t=cmd") == "'\t=cmd"

    @pytest.mark.unit
    def test_sanitize_safe_text_unchanged(self):
        """טקסט רגיל לא משתנה"""
        assert _sanitize_text("משה כהן") == "משה כהן"

    @pytest.mark.unit
    def test_sanitize_empty_string(self):
        """מחרוזת ריקה לא משתנה"""
        assert _sanitize_text("") == ""

    @pytest.mark.unit
    def test_sanitize_non_string(self):
        """ערך שאינו מחרוזת מוחזר כמו שהוא"""
        assert _sanitize_text(123) == 123  # type: ignore[arg-type]


class TestCollectionReportExcel:
    """בדיקות ייצוא דוח גבייה ל-Excel"""

    @pytest.mark.unit
    def test_generates_valid_xlsx(self):
        """ייצוא דוח גבייה מייצר קובץ XLSX תקין"""
        items = [
            {"driver_name": "משה כהן", "total_debt": 300.0, "charge_count": 2},
            {"driver_name": "דני לוי", "total_debt": 200.0, "charge_count": 1},
        ]
        result = generate_collection_report_excel(
            items=items,
            total_debt=500.0,
            cycle_start="2026-01-28",
            cycle_end="2026-02-28",
            station_name="תחנת מבחן",
        )
        assert isinstance(result, bytes)
        assert len(result) > 0

        # וידוא שהקובץ נפתח כ-XLSX תקין
        wb = load_workbook(io.BytesIO(result))
        assert "דוח גבייה" in wb.sheetnames

    @pytest.mark.unit
    def test_contains_data_rows(self):
        """הקובץ מכיל את הנתונים הנכונים"""
        items = [
            {"driver_name": "משה כהן", "total_debt": 300.0, "charge_count": 2},
        ]
        result = generate_collection_report_excel(
            items=items,
            total_debt=300.0,
            cycle_start="2026-01-28",
            cycle_end="2026-02-28",
        )
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        # חיפוש שם הנהג בגליון
        found = False
        for row in ws.iter_rows(values_only=True):
            if "משה כהן" in (row[0] or ""):
                found = True
                assert row[1] == 300.0
                assert row[2] == 2
        assert found, "לא נמצא שם הנהג בגליון"

    @pytest.mark.unit
    def test_rtl_direction(self):
        """הגליון מוגדר כ-RTL"""
        result = generate_collection_report_excel(
            items=[], total_debt=0, cycle_start="2026-01-01", cycle_end="2026-02-01"
        )
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.sheet_view.rightToLeft is True

    @pytest.mark.unit
    def test_empty_items(self):
        """דוח ריק — הקובץ נוצר ללא שגיאה"""
        result = generate_collection_report_excel(
            items=[], total_debt=0, cycle_start="2026-01-01", cycle_end="2026-02-01"
        )
        assert isinstance(result, bytes)
        wb = load_workbook(io.BytesIO(result))
        assert wb.active is not None

    @pytest.mark.unit
    def test_formula_injection_in_driver_name(self):
        """שם נהג שמתחיל ב-= עובר סניטציה בקובץ"""
        items = [
            {"driver_name": "=SUM(A1:A10)", "total_debt": 100.0, "charge_count": 1},
        ]
        result = generate_collection_report_excel(
            items=items, total_debt=100.0,
            cycle_start="2026-01-01", cycle_end="2026-02-01",
        )
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        # וידוא שהערך בתא מתחיל בגרש ולא כנוסחה
        for row in ws.iter_rows(values_only=True):
            if row[0] and "SUM" in str(row[0]):
                assert str(row[0]).startswith("'"), "שם נהג עם = לא עבר סניטציה"


class TestRevenueReportExcel:
    """בדיקות ייצוא דוח הכנסות ל-Excel"""

    @pytest.mark.unit
    def test_generates_valid_xlsx(self):
        """ייצוא דוח הכנסות מייצר קובץ XLSX תקין"""
        result = generate_revenue_report_excel(
            total_commissions=500.0,
            total_manual_charges=200.0,
            total_withdrawals=100.0,
            net_total=600.0,
            date_from="2026-01-01",
            date_to="2026-01-31",
            station_name="תחנת מבחן",
        )
        assert isinstance(result, bytes)

        wb = load_workbook(io.BytesIO(result))
        assert "דוח הכנסות" in wb.sheetnames

    @pytest.mark.unit
    def test_contains_financial_data(self):
        """הקובץ מכיל את הנתונים הפיננסיים"""
        result = generate_revenue_report_excel(
            total_commissions=500.0,
            total_manual_charges=200.0,
            total_withdrawals=100.0,
            net_total=600.0,
            date_from="2026-01-01",
            date_to="2026-01-31",
        )
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        # חיפוש ערך הנטו
        found_net = False
        for row in ws.iter_rows(values_only=True):
            if row[1] == 600.0:
                found_net = True
        assert found_net, "לא נמצא ערך נטו בגליון"


class TestProfitLossExcel:
    """בדיקות ייצוא דוח רווח/הפסד ל-Excel"""

    @pytest.mark.unit
    def test_generates_valid_xlsx(self):
        """ייצוא דוח רווח/הפסד מייצר קובץ XLSX תקין"""
        revenue_by_month = [
            {"month": "2026-01", "commissions": 500.0, "manual_charges": 200.0, "withdrawals": 100.0, "net": 600.0},
            {"month": "2026-02", "commissions": 600.0, "manual_charges": 150.0, "withdrawals": 50.0, "net": 700.0},
        ]
        result = generate_profit_loss_excel(
            revenue_by_month=revenue_by_month,
            date_from="2026-01-01",
            date_to="2026-02-28",
            station_name="תחנת מבחן",
        )
        assert isinstance(result, bytes)

        wb = load_workbook(io.BytesIO(result))
        assert "רווח והפסד" in wb.sheetnames

    @pytest.mark.unit
    def test_monthly_rows(self):
        """הקובץ מכיל שורה לכל חודש"""
        revenue_by_month = [
            {"month": "2026-01", "commissions": 500.0, "manual_charges": 200.0, "withdrawals": 100.0, "net": 600.0},
            {"month": "2026-02", "commissions": 600.0, "manual_charges": 150.0, "withdrawals": 50.0, "net": 700.0},
        ]
        result = generate_profit_loss_excel(
            revenue_by_month=revenue_by_month,
            date_from="2026-01-01",
            date_to="2026-02-28",
        )
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        months_found = []
        for row in ws.iter_rows(values_only=True):
            if row[0] in ("2026-01", "2026-02"):
                months_found.append(row[0])
        assert "2026-01" in months_found
        assert "2026-02" in months_found

    @pytest.mark.unit
    def test_empty_months(self):
        """דוח ריק — ללא חודשים"""
        result = generate_profit_loss_excel(
            revenue_by_month=[], date_from="2026-01-01", date_to="2026-01-31"
        )
        assert isinstance(result, bytes)


class TestMonthlySummaryExcel:
    """בדיקות ייצוא דוח חודשי מסכם ל-Excel"""

    @pytest.mark.unit
    def test_generates_multi_sheet_xlsx(self):
        """דוח חודשי מייצר קובץ עם מספר גליונות"""
        result = generate_monthly_summary_excel(
            month="2026-01",
            station_name="תחנת מבחן",
            collection_items=[
                {"driver_name": "משה כהן", "total_debt": 300.0, "charge_count": 2},
            ],
            total_debt=300.0,
            revenue_data={"commissions": 500.0, "manual_charges": 200.0, "withdrawals": 100.0, "net": 600.0},
            delivery_stats={"total": 50, "delivered": 40, "cancelled": 5, "open": 5},
        )
        assert isinstance(result, bytes)

        wb = load_workbook(io.BytesIO(result))
        assert "סיכום" in wb.sheetnames
        assert "גבייה" in wb.sheetnames

    @pytest.mark.unit
    def test_summary_sheet_contains_stats(self):
        """גליון הסיכום מכיל סטטיסטיקות"""
        result = generate_monthly_summary_excel(
            month="2026-01",
            station_name="תחנת מבחן",
            collection_items=[],
            total_debt=0,
            revenue_data={"commissions": 0, "manual_charges": 0, "withdrawals": 0, "net": 0},
            delivery_stats={"total": 10, "delivered": 8, "cancelled": 1, "open": 1},
        )
        wb = load_workbook(io.BytesIO(result))
        ws = wb["סיכום"]

        # בדיקה שהגליון מכיל כותרת עם שם התחנה
        found_title = False
        for row in ws.iter_rows(values_only=True):
            for cell_value in row:
                if cell_value and "תחנת מבחן" in str(cell_value):
                    found_title = True
        assert found_title, "לא נמצא שם התחנה בגליון הסיכום"
