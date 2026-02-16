"""
בדיקות דוחות מורחבים בפאנל — סעיף 7: ייצוא Excel, דוח רווח/הפסד, דוח חודשי
"""
import pytest
from openpyxl import load_workbook
import io

from app.core.auth import create_access_token
from app.db.models.user import UserRole
from app.db.models.station import Station
from app.db.models.station_wallet import StationWallet
from app.db.models.manual_charge import ManualCharge
from app.db.models.delivery import Delivery, DeliveryStatus
from app.db.models.station_ledger import StationLedger, StationLedgerEntryType


class TestExtendedReports:
    """בדיקות דוחות מורחבים"""

    async def _setup(self, user_factory, db_session, with_data: bool = False):
        owner = await user_factory(
            phone_number="+972501234567",
            name="בעל תחנה",
            role=UserRole.STATION_OWNER,
        )
        station = Station(name="תחנת מבחן", owner_id=owner.id)
        db_session.add(station)
        await db_session.flush()
        wallet = StationWallet(station_id=station.id)
        db_session.add(wallet)

        dispatcher = await user_factory(
            phone_number="+972509999999",
            name="סדרן",
            role=UserRole.COURIER,
        )

        if with_data:
            # חיובים ידניים
            for i, name in enumerate(["משה כהן", "משה כהן", "דני לוי"]):
                charge = ManualCharge(
                    station_id=station.id,
                    dispatcher_id=dispatcher.id,
                    driver_name=name,
                    amount=100.0 * (i + 1),
                    description=f"חיוב {i + 1}",
                )
                db_session.add(charge)

            # תנועות לדג'ר
            for et, amount in [
                (StationLedgerEntryType.COMMISSION_CREDIT, 50.0),
                (StationLedgerEntryType.MANUAL_CHARGE, 200.0),
                (StationLedgerEntryType.WITHDRAWAL, 30.0),
            ]:
                entry = StationLedger(
                    station_id=station.id,
                    entry_type=et,
                    amount=amount,
                    balance_after=amount,
                    description="דוח",
                )
                db_session.add(entry)

            # משלוחים
            sender = await user_factory(
                phone_number="+972501111111",
                name="שולח",
                role=UserRole.SENDER,
            )
            for s in [DeliveryStatus.DELIVERED, DeliveryStatus.DELIVERED, DeliveryStatus.CANCELLED, DeliveryStatus.OPEN]:
                delivery = Delivery(
                    sender_id=sender.id,
                    station_id=station.id,
                    pickup_address="רחוב הרצל 1, תל אביב",
                    dropoff_address="רחוב בן יהודה 50, ירושלים",
                    status=s,
                )
                db_session.add(delivery)

        await db_session.commit()
        return owner, station

    def _get_token(self, user_id: int, station_id: int) -> str:
        return create_access_token(user_id, station_id, "station_owner")

    # ==================== ייצוא Excel — דוח גבייה ====================

    @pytest.mark.asyncio
    async def test_export_collection_xlsx(
        self, test_client, user_factory, db_session,
    ):
        """ייצוא דוח גבייה ל-Excel — headers ותוכן תקינים"""
        owner, station = await self._setup(user_factory, db_session, with_data=True)
        token = self._get_token(owner.id, station.id)

        response = await test_client.get(
            "/api/panel/reports/collection/export-xlsx",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]
        assert "attachment" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]

        # וידוא שהקובץ נפתח כ-XLSX תקין
        wb = load_workbook(io.BytesIO(response.content))
        assert wb.active is not None

    @pytest.mark.asyncio
    async def test_export_collection_xlsx_empty(
        self, test_client, user_factory, db_session,
    ):
        """ייצוא דוח גבייה ריק ל-Excel"""
        owner, station = await self._setup(user_factory, db_session, with_data=False)
        token = self._get_token(owner.id, station.id)

        response = await test_client.get(
            "/api/panel/reports/collection/export-xlsx",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert response.status_code == 200

    # ==================== ייצוא Excel — דוח הכנסות ====================

    @pytest.mark.asyncio
    async def test_export_revenue_xlsx(
        self, test_client, user_factory, db_session,
    ):
        """ייצוא דוח הכנסות ל-Excel"""
        owner, station = await self._setup(user_factory, db_session, with_data=True)
        token = self._get_token(owner.id, station.id)

        response = await test_client.get(
            "/api/panel/reports/revenue/export-xlsx?date_from=2020-01-01&date_to=2030-12-31",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]

        wb = load_workbook(io.BytesIO(response.content))
        assert "דוח הכנסות" in wb.sheetnames

    # ==================== דוח רווח/הפסד ====================

    @pytest.mark.asyncio
    async def test_profit_loss_report(
        self, test_client, user_factory, db_session,
    ):
        """דוח רווח/הפסד מחזיר נתונים"""
        owner, station = await self._setup(user_factory, db_session, with_data=True)
        token = self._get_token(owner.id, station.id)

        response = await test_client.get(
            "/api/panel/reports/profit-loss?date_from=2020-01-01&date_to=2030-12-31",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert response.status_code == 200
        data = response.json()
        assert "months" in data
        assert "total_net" in data
        assert data["date_from"] == "2020-01-01"
        assert data["date_to"] == "2030-12-31"

    @pytest.mark.asyncio
    async def test_profit_loss_report_default_dates(
        self, test_client, user_factory, db_session,
    ):
        """דוח רווח/הפסד עם ברירת מחדל"""
        owner, station = await self._setup(user_factory, db_session)
        token = self._get_token(owner.id, station.id)

        response = await test_client.get(
            "/api/panel/reports/profit-loss",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["months"] == []
        assert data["total_net"] == 0

    @pytest.mark.asyncio
    async def test_profit_loss_export_xlsx(
        self, test_client, user_factory, db_session,
    ):
        """ייצוא דוח רווח/הפסד ל-Excel"""
        owner, station = await self._setup(user_factory, db_session, with_data=True)
        token = self._get_token(owner.id, station.id)

        response = await test_client.get(
            "/api/panel/reports/profit-loss/export-xlsx?date_from=2020-01-01&date_to=2030-12-31",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]

    # ==================== דוח חודשי מסכם ====================

    @pytest.mark.asyncio
    async def test_monthly_summary(
        self, test_client, user_factory, db_session,
    ):
        """דוח חודשי מסכם מחזיר נתונים"""
        owner, station = await self._setup(user_factory, db_session, with_data=True)
        token = self._get_token(owner.id, station.id)

        response = await test_client.get(
            "/api/panel/reports/monthly-summary?month=2026-01",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["month"] == "2026-01"
        assert data["station_name"] == "תחנת מבחן"
        assert "commissions" in data
        assert "total_deliveries" in data
        assert "total_debt" in data

    @pytest.mark.asyncio
    async def test_monthly_summary_default_month(
        self, test_client, user_factory, db_session,
    ):
        """דוח חודשי עם ברירת מחדל (חודש קודם)"""
        owner, station = await self._setup(user_factory, db_session)
        token = self._get_token(owner.id, station.id)

        response = await test_client.get(
            "/api/panel/reports/monthly-summary",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert response.status_code == 200
        data = response.json()
        assert "month" in data

    @pytest.mark.asyncio
    async def test_monthly_summary_invalid_format(
        self, test_client, user_factory, db_session,
    ):
        """דוח חודשי עם פורמט שגוי"""
        owner, station = await self._setup(user_factory, db_session)
        token = self._get_token(owner.id, station.id)

        response = await test_client.get(
            "/api/panel/reports/monthly-summary?month=invalid",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert response.status_code == 400

    @pytest.mark.asyncio
    async def test_monthly_summary_export_xlsx(
        self, test_client, user_factory, db_session,
    ):
        """ייצוא דוח חודשי ל-Excel — מספר גליונות"""
        owner, station = await self._setup(user_factory, db_session, with_data=True)
        token = self._get_token(owner.id, station.id)

        response = await test_client.get(
            "/api/panel/reports/monthly-summary/export-xlsx?month=2026-01",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]

        # בדיקה שיש 2 גליונות
        wb = load_workbook(io.BytesIO(response.content))
        assert len(wb.sheetnames) == 2

    # ==================== אימות — unauthorized ====================

    @pytest.mark.asyncio
    async def test_unauthorized_access(
        self, test_client,
    ):
        """גישה ללא טוקן — 403"""
        endpoints = [
            "/api/panel/reports/collection/export-xlsx",
            "/api/panel/reports/revenue/export-xlsx",
            "/api/panel/reports/profit-loss",
            "/api/panel/reports/profit-loss/export-xlsx",
            "/api/panel/reports/monthly-summary",
            "/api/panel/reports/monthly-summary/export-xlsx",
        ]
        for endpoint in endpoints:
            response = await test_client.get(endpoint)
            assert response.status_code == 403, f"Expected 403 for {endpoint}, got {response.status_code}"
